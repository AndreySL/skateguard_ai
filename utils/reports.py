import os
from datetime import datetime
import pandas as pd
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping
from config import RESULTS_DIR

# Регистрация шрифтов с поддержкой кириллицы
try:
    # Попытка зарегистрировать шрифт DejaVu (часто доступен в системах Linux)
    pdfmetrics.registerFont(TTFont('DejaVu', 'DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVu-Bold', 'DejaVuSans-Bold.ttf'))
    FONT_NAME = 'DejaVu'
except:
    try:
        # Для Windows
        pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
        FONT_NAME = 'Arial'
    except:
        try:
            # Для macOS
            pdfmetrics.registerFont(TTFont('Helvetica', 'Helvetica'))
            FONT_NAME = 'Helvetica'
        except:
            # Fallback - используем стандартный шрифт (может не поддерживать кириллицу)
            FONT_NAME = 'Helvetica'

def generate_pdf_report(history: list) -> str:
    """Генерация PDF-отчёта с поддержкой русского языка"""
    pdf_path = os.path.join(RESULTS_DIR, f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    
    # Создаём документ в альбомной ориентации для лучшего отображения
    doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    # Создаём стили с поддержкой кириллицы
    styles = getSampleStyleSheet()
    
    # Стиль для заголовка
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontName=FONT_NAME,
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=30,
        textColor=colors.HexColor('#1a237e')
    )
    
    # Стиль для обычного текста
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=10,
        alignment=TA_LEFT,
        spaceAfter=12
    )
    
    # Стиль для заголовков таблицы
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.white,
        bold=True
    )
    
    # Стиль для ячеек таблицы
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=8,
        alignment=TA_LEFT
    )
    
    story = []
    
    # Заголовок отчёта
    story.append(Paragraph("SkateGuard AI - Отчёт по контролю катания", title_style))
    story.append(Spacer(1, 10))
    
    # Дата генерации
    date_str = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
    story.append(Paragraph(f"Дата формирования отчёта: {date_str}", normal_style))
    story.append(Spacer(1, 20))
    
    # Статистика
    total = len(history)
    violations = sum(1 for h in history if h.get("violation"))
    violation_percent = (violations / total * 100) if total > 0 else 0
    
    stats_text = f"""
    <b>Общая статистика:</b><br/>
    • Всего анализов: {total}<br/>
    • Нарушений: {violations}<br/>
    • Процент нарушений: {violation_percent:.1f}%<br/>
    • Без нарушений: {total - violations}
    """
    story.append(Paragraph(stats_text, normal_style))
    story.append(Spacer(1, 20))
    
    # Заголовки таблицы (на русском)
    headers = ['Время', 'Тип', 'Имя файла', 'Местность', 'Нарушение', 'Причина нарушения']
    
    # Подготовка данных для таблицы
    data = [headers]
    
    for h in history[:50]:  # Ограничиваем 50 записями для читаемости
        # Форматирование времени
        timestamp = h.get("timestamp", "")
        if timestamp:
            try:
                dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                timestamp = dt.strftime('%d.%m.%Y %H:%M:%S')
            except:
                pass
        
        # Тип анализа
        analysis_type = "📷 Изображение" if h.get("type") == "image" else "🎬 Видео"
        
        # Имя файла
        filename = h.get("name", "Неизвестно")[:25]
        
        # Местность
        location = h.get("location_display_text", h.get("location_type", "Не определено"))
        if location and len(location) > 30:
            location = location[:27] + "..."
        elif not location:
            location = "Не определено"
        
        # Статус нарушения
        if h.get("violation"):
            violation_status = "⚠️ ДА"
            violation_status_style = 'danger'
        else:
            violation_status = "✅ НЕТ"
            violation_status_style = 'success'
        
        # Причина нарушения
        violation_reason = h.get("violation_reason", "")
        if violation_reason and len(violation_reason) > 40:
            violation_reason = violation_reason[:37] + "..."
        elif not violation_reason:
            violation_reason = "-"
        
        data.append([
            timestamp,
            analysis_type,
            filename,
            location,
            violation_status,
            violation_reason
        ])
    
    # Создание таблицы
    # Автоматическое определение ширины колонок
    col_widths = [100, 90, 120, 100, 70, 150]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Стилизация таблицы
    table_style = TableStyle([
        # Заголовок
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOLD', (0, 0), (-1, 0), True),
        
        # Ячейки
        ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Время по центру
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Тип по центру
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Нарушение по центру
        
        # Границы
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Чередование цветов строк
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        
        # Выравнивание по вертикали
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        
        # Отступы в ячейках
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ])
    
    # Подсветка строк с нарушениями
    for i, row in enumerate(data[1:], start=1):
        if row[4] == "⚠️ ДА":
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#ffebee'))
            table_style.add('TEXTCOLOR', (0, i), (-1, i), colors.HexColor('#c62828'))
    
    table.setStyle(table_style)
    story.append(table)
    
    # Добавление информации о фильтрах (если есть)
    if len(history) > 50:
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"<i>Примечание: показаны последние 50 записей из {len(history)}</i>", normal_style))
    
    # Добавление нижнего колонтитула
    story.append(Spacer(1, 30))
    footer_text = f"Отчёт сгенерирован системой SkateGuard AI • {datetime.now().strftime('%d.%m.%Y')}"
    story.append(Paragraph(f"<font size=8>{footer_text}</font>", normal_style))
    
    # Сборка документа
    doc.build(story)
    return pdf_path


def generate_excel_report(history: list) -> str:
    """Генерация Excel-отчёта с поддержкой русского языка"""
    excel_path = os.path.join(RESULTS_DIR, f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    # Создаём DataFrame
    df = pd.DataFrame(history)
    
    # Переименовываем колонки на русские
    column_names = {
        'timestamp': 'Время',
        'type': 'Тип анализа',
        'name': 'Имя файла',
        'location_type': 'Тип местности',
        'location_display_text': 'Местность',
        'violation': 'Нарушение',
        'violation_reason': 'Причина нарушения',
        'skateboards_detected': 'Обнаружено скейтбордов',
        'persons_detected': 'Обнаружено людей',
        'confidence': 'Уверенность (%)',
        'duration': 'Длительность (сек)',
        'total_frames': 'Всего кадров',
        'violation_percentage': 'Процент нарушений (%)',
        'avg_skateboards': 'Среднее скейтбордов'
    }
    
    # Применяем переименование только для существующих колонок
    existing_columns = {k: v for k, v in column_names.items() if k in df.columns}
    df = df.rename(columns=existing_columns)
    
    # Форматирование времени
    if 'Время' in df.columns:
        df['Время'] = pd.to_datetime(df['Время'], errors='coerce')
        df['Время'] = df['Время'].dt.strftime('%d.%m.%Y %H:%M:%S')
    
    # Замена значений на русские
    if 'Тип анализа' in df.columns:
        df['Тип анализа'] = df['Тип анализа'].map({'image': '📷 Изображение', 'video': '🎬 Видео'})
    
    if 'Нарушение' in df.columns:
        df['Нарушение'] = df['Нарушение'].map({True: '⚠️ ДА', False: '✅ НЕТ'})
    
    # Заполнение пустых значений
    df = df.fillna('-')
    
    # Создание Excel файла с поддержкой UTF-8
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Отчёт SkateGuard', index=False)
        
        # Получаем workbook и worksheet для форматирования
        workbook = writer.book
        worksheet = writer.sheets['Отчёт SkateGuard']
        
        # Настройка ширины колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Стилизация заголовков
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Выравнивание ячеек
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Подсветка строк с нарушениями
                violation_col = None
                for idx, col in enumerate(worksheet[1], 1):
                    if col.value == 'Нарушение':
                        violation_col = idx
                        break
                
                if violation_col and cell.column == violation_col and cell.value == '⚠️ ДА':
                    cell.font = Font(color='C62828', bold=True)
                    # Подсветка всей строки
                    for c in row:
                        c.font = Font(color='C62828')
    
    return excel_path